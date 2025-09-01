import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import FormulaRule
from datetime import datetime, time, timedelta
import urllib.parse
import json

class UNEGarioCalendar:
    def __init__(self):
        # Fechas del semestre Ago-Dic 2025
        self.fecha_inicio = "20250818"
        self.fecha_fin = "20251212"
        
        # Días festivos y recesos según calendario académico ISEC 2025-2026
        self.dias_festivos = [
            # Agosto 2025
            "20250825", "20250831",  # 25 y 31 de agosto
            # Septiembre 2025  
            "20250916",  # 16 de septiembre (Independencia)
            "20250929", "20250930",  # 29 y 30 de septiembre
            # Octubre 2025
            "20251010",  # 10 de octubre
            # Noviembre 2025
            "20251102", "20251117",  # 2 y 17 de noviembre
            # Diciembre 2025
            "20251208",  # 8 de diciembre
            "20251222", "20251223", "20251224", "20251225", "20251226", "20251229", "20251230", "20251231",
            # Enero 2026
            "20260101", "20260102", "20260105", "20260106", "20260112", "20260119", "20260126",
            # Febrero 2026  
            "20260203", "20260216"  # 3 y 16 de febrero
        ]
        
        # Mapeo de días español a código
        self.dias_map = {
            "Lunes": "MO",
            "Martes": "TU", 
            "Miércoles": "WE",
            "Jueves": "TH",
            "Viernes": "FR",
            "Sábado": "SA",
            "Domingo": "SU"
        }
        
        # Colores para materias según tu horario real
        self.colores = {
            "Cálculo multivariable": {"hex": "FFD966", "google_id": "5"},
            "Finanzas empresariales": {"hex": "9BC2E6", "google_id": "7"},
            "Algoritmos y estructuras de datos": {"hex": "A9D08E", "google_id": "2"},
            "Fundamentos de diseño digital": {"hex": "F4B183", "google_id": "6"},
            "Álgebra lineal": {"hex": "C6E0B4", "google_id": "10"},
            "Ingeniería, ética y sociedad": {"hex": "D9D2E9", "google_id": "3"},
            "Receso": {"hex": "E7E6E6", "google_id": "8"}
        }

    def leer_horario_excel(self, archivo="UNEGario.xlsx"):
        """Lee el archivo Excel y extrae los datos del horario"""
        wb = openpyxl.load_workbook(archivo)
        ws = wb["Horario"]
        
        horario = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:  # Si hay día
                horario.append({
                    "dia": row[0],
                    "hora_inicio": row[1],
                    "hora_fin": row[2],
                    "materia": row[3],
                    "profesor": row[4] or "",
                    "tipo": "Receso" if row[3] == "Receso" else "Clase"
                })
        
        wb.close()
        return horario

    def generar_datos_horario(self):
        """Genera los datos del horario completo con el formato correcto basado en tu horario real"""
        # Horario completo de la semana EXACTO según tu imagen
        horario_completo = [
            # Lunes
            {"dia": "Lunes", "hora_inicio": time(17,0), "hora_fin": time(18,0), 
             "materia": "Cálculo multivariable", "profesor": "ORTIZ REYES RAUL", "tipo": "Clase"},
            {"dia": "Lunes", "hora_inicio": time(20,0), "hora_fin": time(21,30), 
             "materia": "Finanzas empresariales", "profesor": "ARROYO SANCHEZ OSCAR RAYMUNDO", "tipo": "Clase"},
            
            # Martes
            {"dia": "Martes", "hora_inicio": time(17,30), "hora_fin": time(18,0), 
             "materia": "Algoritmos y estructuras de datos", "profesor": "PEÑA OLIVO CRISTÓBAL", "tipo": "Clase"},
            {"dia": "Martes", "hora_inicio": time(18,0), "hora_fin": time(19,0), 
             "materia": "Álgebra lineal", "profesor": "LARA USCANGA SERGIO TULIO", "tipo": "Clase"},
            {"dia": "Martes", "hora_inicio": time(19,0), "hora_fin": time(20,0), 
             "materia": "Fundamentos de diseño digital", "profesor": "AGUILAR VELA ÁLVARO", "tipo": "Clase"},
            {"dia": "Martes", "hora_inicio": time(20,0), "hora_fin": time(21,0), 
             "materia": "Finanzas empresariales", "profesor": "ARROYO SANCHEZ OSCAR RAYMUNDO", "tipo": "Clase"},
            {"dia": "Martes", "hora_inicio": time(21,0), "hora_fin": time(21,30), 
             "materia": "Fundamentos de diseño digital", "profesor": "AGUILAR VELA ÁLVARO", "tipo": "Clase"},
            
            # Miércoles
            {"dia": "Miércoles", "hora_inicio": time(17,30), "hora_fin": time(18,0), 
             "materia": "Cálculo multivariable", "profesor": "ORTIZ REYES RAUL", "tipo": "Clase"},
            {"dia": "Miércoles", "hora_inicio": time(18,0), "hora_fin": time(19,0), 
             "materia": "Álgebra lineal", "profesor": "LARA USCANGA SERGIO TULIO", "tipo": "Clase"},
            {"dia": "Miércoles", "hora_inicio": time(19,0), "hora_fin": time(20,0), 
             "materia": "Fundamentos de diseño digital", "profesor": "AGUILAR VELA ÁLVARO", "tipo": "Clase"},
            {"dia": "Miércoles", "hora_inicio": time(20,0), "hora_fin": time(21,0), 
             "materia": "Finanzas empresariales", "profesor": "ARROYO SANCHEZ OSCAR RAYMUNDO", "tipo": "Clase"},
            {"dia": "Miércoles", "hora_inicio": time(21,0), "hora_fin": time(21,30), 
             "materia": "Finanzas empresariales", "profesor": "ARROYO SANCHEZ OSCAR RAYMUNDO", "tipo": "Clase"},
            
            # Jueves
            {"dia": "Jueves", "hora_inicio": time(19,0), "hora_fin": time(20,0), 
             "materia": "Fundamentos de diseño digital", "profesor": "AGUILAR VELA ÁLVARO", "tipo": "Clase"},
            {"dia": "Jueves", "hora_inicio": time(20,0), "hora_fin": time(21,0), 
             "materia": "Ingeniería, ética y sociedad", "profesor": "CAMPOAMOR ROLDAN SERGIO", "tipo": "Clase"},
            {"dia": "Jueves", "hora_inicio": time(21,0), "hora_fin": time(21,30), 
             "materia": "Finanzas empresariales", "profesor": "ARROYO SANCHEZ OSCAR RAYMUNDO", "tipo": "Clase"},
            
            # Viernes
            {"dia": "Viernes", "hora_inicio": time(17,0), "hora_fin": time(18,0), 
             "materia": "Álgebra lineal", "profesor": "LARA USCANGA SERGIO TULIO", "tipo": "Clase"},
            {"dia": "Viernes", "hora_inicio": time(20,0), "hora_fin": time(21,0), 
             "materia": "Ingeniería, ética y sociedad", "profesor": "CAMPOAMOR ROLDAN SERGIO", "tipo": "Clase"},
            {"dia": "Viernes", "hora_inicio": time(21,0), "hora_fin": time(21,30), 
             "materia": "Ingeniería, ética y sociedad", "profesor": "CAMPOAMOR ROLDAN SERGIO", "tipo": "Clase"},
            
            # Sábado
            {"dia": "Sábado", "hora_inicio": time(7,0), "hora_fin": time(9,0), 
             "materia": "Algoritmos y estructuras de datos", "profesor": "PEÑA OLIVO CRISTÓBAL", "tipo": "Clase"},
            {"dia": "Sábado", "hora_inicio": time(9,0), "hora_fin": time(10,0), 
             "materia": "Algoritmos y estructuras de datos", "profesor": "PEÑA OLIVO CRISTÓBAL", "tipo": "Clase"},
            {"dia": "Sábado", "hora_inicio": time(10,0), "hora_fin": time(11,0), 
             "materia": "Ingeniería, ética y sociedad", "profesor": "CAMPOAMOR ROLDAN SERGIO", "tipo": "Clase"},
            {"dia": "Sábado", "hora_inicio": time(11,0), "hora_fin": time(12,0), 
             "materia": "Ingeniería, ética y sociedad", "profesor": "CAMPOAMOR ROLDAN SERGIO", "tipo": "Clase"},
            {"dia": "Sábado", "hora_inicio": time(12,0), "hora_fin": time(13,0), 
             "materia": "Ingeniería, ética y sociedad", "profesor": "CAMPOAMOR ROLDAN SERGIO", "tipo": "Clase"},
        ]
        
        return horario_completo

    def generar_rrule(self):
        """Genera la regla de recurrencia RRULE"""
        # Obtener todos los días únicos del horario
        horario = self.generar_datos_horario()
        dias_unicos = set()
        for clase in horario:
            if clase["tipo"] == "Clase":
                dias_unicos.add(self.dias_map.get(clase["dia"], ""))
        
        dias_str = ",".join(sorted(dias_unicos, key=lambda x: ["MO","TU","WE","TH","FR","SA","SU"].index(x)))
        
        rrule = f"FREQ=WEEKLY;BYDAY={dias_str};UNTIL={self.fecha_fin}T235959Z"
        return rrule

    def generar_exdates(self):
        """Genera las fechas de exclusión EXDATE"""
        exdates = []
        for fecha in self.dias_festivos:
            exdates.append(f"{fecha}T000000Z")
        
        return ",".join(exdates)

    def generar_descripcion_html(self):
        """Genera una descripción HTML del horario"""
        horario = self.generar_datos_horario()
        
        descripcion = []
        descripcion.append("📚 HORARIO SEMESTRE AGO-DIC 2025")
        descripcion.append("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
        descripcion.append("")
        
        dia_actual = ""
        for clase in horario:
            if clase["dia"] != dia_actual:
                if dia_actual:
                    descripcion.append("")
                descripcion.append(f"📅 {clase['dia'].upper()}")
                dia_actual = clase["dia"]
            
            hora_inicio = clase["hora_inicio"].strftime("%H:%M")
            hora_fin = clase["hora_fin"].strftime("%H:%M")
            
            if clase["tipo"] == "Clase":
                descripcion.append(f"  • {hora_inicio}-{hora_fin} | {clase['materia']}")
                if clase["profesor"]:
                    descripcion.append(f"    Profesor: {clase['profesor']}")
            else:
                descripcion.append(f"  • {hora_inicio}-{hora_fin} | ☕ RECESO")
        
        descripcion.append("")
        descripcion.append("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
        descripcion.append("🔗 Más información: https://edbetosolutions.tech/UNEGarios")
        descripcion.append("")
        descripcion.append("⚠️ DÍAS SIN CLASES:")
        descripcion.append("• 16 sep - Día de la Independencia")
        descripcion.append("• 2 nov - Día de Muertos")
        descripcion.append("• 17 nov - Revolución Mexicana")
        descripcion.append("• 8 dic - Inmaculada Concepción")
        descripcion.append("• 22 dic - 6 ene - Vacaciones de Invierno")
        
        return "%0A".join(descripcion)

    def generar_url_google_calendar(self):
        """Genera la URL completa para añadir a Google Calendar"""
        # Parámetros básicos
        titulo = "Horario UNEGario - Semestre Ago-Dic 2025"
        fecha_inicio_evento = f"{self.fecha_inicio}T170000Z"  # 17:00 primer día
        fecha_fin_evento = f"{self.fecha_inicio}T220000Z"     # 22:00 primer día
        ubicacion = "Universidad ISEC - Campus Principal"
        
        # Generar RRULE y EXDATE
        rrule = self.generar_rrule()
        exdates = self.generar_exdates()
        
        # Generar descripción
        descripcion = self.generar_descripcion_html()
        
        # Construir recurrencia completa
        recur = f"RRULE:{rrule}"
        if exdates:
            recur += f"%0AEXDATE;VALUE=DATE:{exdates.replace(',', '%2C')}"
        
        # Construir URL
        base_url = "https://calendar.google.com/calendar/render"
        params = {
            "action": "TEMPLATE",
            "text": titulo,
            "dates": f"{fecha_inicio_evento}/{fecha_fin_evento}",
            "details": descripcion,
            "location": ubicacion,
            "ctz": "America/Mexico_City",
            "recur": recur
        }
        
        # Codificar parámetros
        param_str = "&".join([f"{k}={urllib.parse.quote(str(v), safe='')}" for k, v in params.items()])
        
        return f"{base_url}?{param_str}"

    def generar_json_horario(self):
        """Genera un JSON con el horario para usar en JavaScript"""
        horario = self.generar_datos_horario()
        
        # Convertir objetos time a strings
        horario_json = []
        for clase in horario:
            horario_json.append({
                "dia": clase["dia"],
                "hora_inicio": clase["hora_inicio"].strftime("%H:%M"),
                "hora_fin": clase["hora_fin"].strftime("%H:%M"),
                "materia": clase["materia"],
                "profesor": clase["profesor"],
                "tipo": clase["tipo"],
                "color": self.colores.get(clase["materia"], {}).get("hex", "FFFFFF")
            })
        
        return json.dumps(horario_json, ensure_ascii=False, indent=2)

    def exportar_archivos_web(self):
        """Exporta todos los archivos necesarios para la web"""
        # Generar URL de Google Calendar
        url_calendar = self.generar_url_google_calendar()
        
        # Generar JSON del horario
        horario_json = self.generar_json_horario()
        
        # Asegurar que existe la carpeta output
        import os
        os.makedirs("../output", exist_ok=True)
        
        # Guardar JSON
        with open("../output/horario_data.json", "w", encoding="utf-8") as f:
            f.write(horario_json)
        
        # Guardar URL en archivo de texto
        with open("../output/google_calendar_url.txt", "w", encoding="utf-8") as f:
            f.write(url_calendar)
        
        print("✅ Archivos generados:")
        print("  - ../output/horario_data.json")
        print("  - ../output/google_calendar_url.txt")
        print(f"\n📅 URL de Google Calendar generada")
        print("   Puedes copiar esta URL para el botón en tu página web")
        
        return url_calendar, horario_json

if __name__ == "__main__":
    unegario = UNEGarioCalendar()
    url, json_data = unegario.exportar_archivos_web()
